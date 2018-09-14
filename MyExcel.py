# time:2018-08-07 14:20
from openpyxl import load_workbook
import logging

class MyExcel:
    def __init__(self,excelFile):
        # 打开excel文件
        self.wb = load_workbook(excelFile)
        # 获取sheet
        self.caseData_Sh = self.wb["casedatas"]
        self.initData_sh = self.wb["params_datas"]
    # 读取初始化数据   当函数名前加两个——表示完全私有化，只有自己的类能看到
    def __get_initDatas(self):
        initDatas = {}
        for row in range(2,self.initData_sh.max_row + 1):
            initDatas[self.initData_sh.cell(row=row,column=1).value] = str(self.initData_sh.cell(row=row,column=2).value)
            # 设置phone、noreg_phone的值
            initDatas["${phone}"] = str(int(initDatas["${init_phone}"]) + 1)
            initDatas["${noreg_phone}"] = str(int(initDatas["${init_phone}"]) + 2)
        logging.info("初始化数据为：\n{0}".format(initDatas))
        return initDatas
    # 读取测试数据  -- 所有的
    def get_all_testDatas(self):
        caseDatas = []
        # 读取初始化数据
        init_datas = self.__get_initDatas()
        for row in range(2,self.caseData_Sh.max_row+1):
            case_data = {}
            case_data["case_id"] =self.caseData_Sh.cell(row=row,column=1).value
            case_data["api_name"] = self.caseData_Sh.cell(row=row, column=4).value
            case_data["method"] = self.caseData_Sh.cell(row=row, column=5).value
            case_data["url"] = self.caseData_Sh.cell(row=row, column=6).value
            case_data["request_data"] = self.caseData_Sh.cell(row=row, column=7).value
            case_data["expected_data"] = self.caseData_Sh.cell(row=row, column=8).value
            # 添加比对方式
            case_data["compare_type"] = self.caseData_Sh.cell(row=row, column=9).value
            # 初始化请求数据：如果有初始化的数据，则替换掉
            for key,value in init_datas.items():
                if case_data["request_data"].find(key) != -1:
                    logging.info("本条测试数据需要初始化处理！！")
                    case_data["request_data"] = case_data["request_data"].replace(key,value)
            logging.info("初始化之后的测试数据为：\n{0}".format(case_data))
            # 本用例是否需要从返回的结果当中提取数据，如果要，那存储提取表达式
            if self.caseData_Sh.cell(row=row, column=10).value is not None:
                case_data["expression"] = self.caseData_Sh.cell(row=row, column=10).value

            caseDatas.append(case_data)
        return caseDatas
    # 更新初始化数据--初始化手机号，需要每次自动化变化
    def updata_initPhone(self):
        self.initData_sh.cell(row =2,column=2).value = str(int(self.initData_sh.cell(row =2,column=2).value) + 4)

    # 保存并关闭excel
    def save_excel(self,filePath):
        self.wb.save(filePath)



